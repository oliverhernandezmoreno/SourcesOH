import csv
import json
import os
from pathlib import Path
import tempfile

from django.utils.dateparse import parse_datetime as django_parse_datetime
import openpyxl
import xlrd


def parse_workbook(f, declared_name=None):
    """Attempt to read the contents of *f* assuming it's an excel file or
    a csv file. Return a list of sheets, each sheet being a dictionary
    with a 'data' key, it being a list of rows, each row being a list
    of values.

    """
    if not declared_name:
        return parse_openpy(f)
    if declared_name.endswith(".xlsx"):
        return parse_openpy(f)
    if declared_name.endswith(".xls"):
        return parse_xlrd(f)
    if declared_name.endswith(".csv"):
        return parse_csv(f)
    if declared_name.endswith(".tsv"):
        return parse_csv(f, dialect="excel-tab")
    return parse_openpy(f)


def parse_openpy(f):
    wb = None
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        return [
            {"name": name, "data": list(sheet.iter_rows(values_only=True))}
            for name, sheet in zip(wb.sheetnames, wb.worksheets)
        ]
    finally:
        if wb:
            wb.close()


def parse_xlrd(f):
    with xlrd.open_workbook(file_contents=f.read()) as wb:
        return [
            {
                "name": sheet.name,
                "data": [
                    tuple([cell.value for cell in row])
                    for row in sheet.get_rows()
                ]
            }
            for sheet in wb.sheets()
        ]


def parse_csv(f, *args, **kwargs):
    fd, name = tempfile.mkstemp()
    try:
        os.close(fd)
        with open(name, "wb") as temp_file:
            for chunk in f.chunks():
                temp_file.write(chunk)
        with open(name, newline="") as temp_file:
            reader = csv.reader(temp_file, *args, **kwargs)
            return [{
                "name": "",
                "data": [
                    tuple(row)
                    for row in reader
                ]
            }]
    finally:
        Path(name).unlink()


def parse_timestamp(value):
    if not isinstance(value, str):
        return None
    try:
        return django_parse_datetime(value)
    except (TypeError, ValueError):
        return None


def parse_numeric(value):
    if isinstance(value, (int, float)):
        return value
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_integer(value):
    if isinstance(value, int):
        return value
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def parse_json(value):
    if not isinstance(value, str):
        return None
    try:
        return json.loads(value)
    except Exception:
        return None


def parse_text(t):
    if t is None:
        return None
    return str(t).strip()
